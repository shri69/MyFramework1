import openpyxl


class HomePageData:
    test_homepage_Data=[{"firstName":"geetu","lastname":"geetushri"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("H:\\PythonDemo.xlsx")
        sheet = book.active


        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(1, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]